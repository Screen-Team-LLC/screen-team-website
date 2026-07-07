#!/usr/bin/env python3
"""Run Serper.dev SERP research for Screen Team LLC — GSC + page-map queries."""

from __future__ import annotations

import json
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

import requests

from serp_query_map import DEFAULT_LOCATION, build_full_page_query_map, query_locations

ROOT = Path(__file__).resolve().parents[1]
SEO = ROOT / "seo"
OUTREACH_ENV = Path(r"E:\KnightLogics-Growth-System\CRM\OutreachEngine\.env")
GSC_XLSX = ROOT / "screenteamllc.com-Performance-on-Search-2026-07-07.xlsx"
SERPER_ENDPOINT = "https://google.serper.dev/search"
SITE_DOMAIN = "screenteamllc.com"
REQUEST_DELAY_SEC = 1.0

PAGE_QUERY_MAP = build_full_page_query_map()


def load_serper_key() -> str:
    if not OUTREACH_ENV.exists():
        raise RuntimeError(f"Serper env not found: {OUTREACH_ENV}")
    for raw_line in OUTREACH_ENV.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        if key.strip() == "SERPER_API_KEY":
            token = value.strip().strip('"').strip("'")
            if token:
                return token
    raise RuntimeError("SERPER_API_KEY missing in OutreachEngine/.env")


def normalize_domain(url: str) -> str:
    try:
        return urlparse(url).netloc.lower().removeprefix("www.")
    except Exception:
        return ""


def load_gsc_queries() -> list[dict[str, str]]:
    if not GSC_XLSX.exists():
        return []
    try:
        import openpyxl
    except ImportError:
        return []
    wb = openpyxl.load_workbook(GSC_XLSX, read_only=True, data_only=True)
    if "Queries" not in wb.sheetnames:
        return []
    ws = wb["Queries"]
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(x) for x in next(rows_iter)]
    out: list[dict[str, str]] = []
    for row in rows_iter:
        if not row or not row[0]:
            continue
        data = dict(zip(headers, row))
        query = str(data.get("Top queries") or "").strip()
        if not query:
            continue
        out.append(
            {
                "query": query,
                "impressions": str(int(data.get("Impressions") or 0)),
                "position": str(data.get("Position") or ""),
            }
        )
    return out


def map_gsc_query_to_page(query: str) -> str:
    q = query.lower()
    rules = [
        ("pool cage repair clearwater", "clearwater-screen-repair.html"),
        ("lanai screen repair seminole", "seminole-screen-repair.html"),
        ("pool screen repair pinellas park", "pinellas-park-screen-repair.html"),
        ("pool cage rescreen gulfport", "gulfport-screen-repair.html"),
        ("pool cage rescreen belleair", "belleair-screen-repair.html"),
        ("screen repair land o lakes", "land-o-lakes-screen-repair.html"),
        ("pinellas county", "pinellas-county-screen-repair.html"),
        ("pasco county", "pasco-county-screen-repair.html"),
        ("hillsborough", "hillsborough-county-screen-repair.html"),
        ("storm screen", "storm-screen-repair.html"),
        ("hurricane screen", "storm-screen-repair.html"),
        ("gutter", "gutters-and-screens.html"),
        ("gutters and screens", "gutters-and-screens.html"),
        ("superscreen", "pet-resistant-screen-mesh.html"),
        ("pet resistant", "pet-resistant-screen-mesh.html"),
        ("lanai screen replacement", "lanai-screen-replacement.html"),
        ("lanai screen repair", "screen-lanais.html"),
        ("pool cage rescreen", "rescreens.html"),
        ("pool cage repair", "pool-cage-repair.html"),
        ("pool rescreen", "rescreens.html"),
        ("screen panel", "screen-panel-repair.html"),
        ("screen repair tampa", "tampa-screen-repair.html"),
        ("screen repair clearwater", "clearwater-screen-repair.html"),
        ("screen team", "index.html"),
    ]
    for needle, page in rules:
        if needle in q:
            return page
    return "index.html"


def collect_queries() -> list[dict[str, Any]]:
    seen: set[str] = set()
    items: list[dict[str, Any]] = []

    def add(query: str, *, page: str, source: str, impressions: str = "", location: str = "") -> None:
        key = f"{query.strip().lower()}|{page}|{location or query_locations(page)}"
        if not query.strip() or key in seen:
            return
        seen.add(key)
        items.append(
            {
                "query": query.strip(),
                "target_page": page,
                "source": source,
                "gsc_impressions": impressions,
                "location": location or query_locations(page),
            }
        )

    for page, queries in PAGE_QUERY_MAP.items():
        location = query_locations(page)
        for query in queries:
            add(query, page=page, source="page_map", location=location)

    for row in load_gsc_queries():
        query = row["query"]
        impressions = int(row["impressions"] or 0)
        if impressions < 10:
            continue
        page = map_gsc_query_to_page(query)
        add(query, page=page, source="gsc", impressions=row["impressions"], location=query_locations(page))

    return items


def fetch_serp(api_key: str, query: str, location: str) -> dict[str, Any]:
    response = requests.post(
        SERPER_ENDPOINT,
        headers={"X-API-KEY": api_key, "Content-Type": "application/json"},
        json={"q": query, "gl": "us", "hl": "en", "num": 10, "location": location},
        timeout=30,
    )
    response.raise_for_status()
    return response.json()


def summarize_organic(data: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for index, item in enumerate(data.get("organic") or [], start=1):
        if not isinstance(item, dict):
            continue
        link = str(item.get("link") or "")
        rows.append(
            {
                "position": index,
                "title": item.get("title"),
                "link": link,
                "domain": normalize_domain(link),
                "snippet": item.get("snippet"),
                "is_screenteam": SITE_DOMAIN in normalize_domain(link),
            }
        )
    return rows


def summarize_places(data: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in data.get("places") or []:
        if not isinstance(item, dict):
            continue
        rows.append(
            {
                "title": item.get("title"),
                "address": item.get("address"),
                "rating": item.get("rating"),
                "ratingCount": item.get("ratingCount"),
                "phone": item.get("phoneNumber"),
                "website": item.get("website"),
            }
        )
    return rows


def infer_intent(query: str, organic: list[dict[str, Any]], places: list[dict[str, Any]]) -> str:
    q = query.lower()
    if places:
        return "local_service_map_pack"
    if any(token in q for token in ("price", "cost", "estimate", "quote", "how much")):
        return "pricing_estimate"
    if "screen team" in q or "chris westcott" in q:
        return "brand_navigational"
    if any(token in q for token in ("near me", "pinellas", "pasco", "hillsborough", "clearwater", "tampa", "seminole")):
        return "local_service"
    domains = {row["domain"] for row in organic[:5]}
    if any("yelp." in d or "angi." in d or "homeadvisor" in d for d in domains):
        return "local_directory_mix"
    return "general_service"


def city_from_page(page: str) -> str:
    city = page.replace("-screen-repair.html", "").replace("-", " ").title()
    if city == "Land O Lakes":
        return "Land O Lakes"
    return city


def suggested_meta_description(row: dict[str, Any]) -> str:
    """SERP-informed meta description for Screen Team (owner-direct, no license claims)."""
    page = str(row.get("target_page") or "")
    intent = str(row.get("intent") or "")
    phone = "Call (727) 386-6562"
    brand = "Chris Westcott on every job"

    # Page-first templates — never inject competitor SERP snippets into meta copy.
    if page == "contact.html":
        return f"{phone} for pool cage, lanai & window screen quotes across Pasco, Pinellas & Hillsborough. {brand}."

    if page == "pricing.html":
        return f"Pool cage rescreen pricing Tampa Bay — panels from $75, full cages from $1,995. $150 minimum. {phone}."

    if page == "index.html" or intent == "brand_navigational":
        return f"Screen repair Tampa Bay — pool cages, lanais & rescreens. {brand}. {phone}."

    if page == "pool-cage-repair.html":
        return f"Pool cage repair Tampa Bay — torn panels, doors & storm damage fixed fast. {brand}. {phone}."

    if page == "rescreens.html":
        return f"Pool cage rescreen Tampa Bay from $1,995. Partial & full jobs. {brand}. {phone}."

    if page == "screen-lanais.html":
        return f"Lanai screen repair Tampa Bay — popped spline, door panels & roof lines. {phone}."

    if page == "lanai-screen-replacement.html":
        return f"Lanai screen replacement Tampa Bay — fresh mesh & aligned doors. {brand}. {phone}."

    if page == "storm-screen-repair.html":
        return f"Storm screen repair Tampa Bay — hurricane panel patches on pool cages & lanais. {phone}."

    if page == "gutters-and-screens.html":
        return f"Gutters & screens Tampa Bay — fix overflow damage on lanais & pool cages together. {phone}."

    if page == "pet-resistant-screen-mesh.html":
        return f"Pet-resistant pool screen mesh Tampa Bay — tougher door panels for dogs. {phone}."

    if page == "screen-panel-repair.html":
        return f"Screen panel repair Tampa Bay from $75/panel. $150 minimum. {brand}. {phone}."

    if page == "pool-enclosures.html":
        return f"Pool enclosure repair Tampa Bay — torn mesh, sagging panels & door tracks. {brand}. {phone}."

    if page == "pinellas-county-screen-repair.html":
        return f"Pinellas County screen repair — pool cages, lanais & enclosures. {brand}. {phone}."

    if page == "pasco-county-screen-repair.html":
        return f"Pasco County screen repair — pool cages, lanais & enclosures. {brand}. {phone}."

    if page == "hillsborough-county-screen-repair.html":
        return f"Hillsborough County screen repair — pool cages, lanais & enclosures. {brand}. {phone}."

    if page.endswith("-screen-repair.html"):
        city = city_from_page(page)
        return f"Pool cage & lanai screen repair in {city}, FL. {brand}. {phone}."

    query = str(row.get("query") or "").lower()
    if "cost" in query or "price" in query:
        return f"Pool cage rescreen pricing Tampa Bay — panels from $75, full cages from $1,995. $150 minimum. {phone}."

    return f"Screen repair Tampa Bay — pool cages, lanais & rescreens. {brand}. {phone}."


def main() -> int:
    api_key = load_serper_key()
    queries = collect_queries()
    generated_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    results: list[dict[str, Any]] = []

    print(f"Running {len(queries)} Serper queries across {len(PAGE_QUERY_MAP)} pages...")
    for index, item in enumerate(queries):
        query = item["query"]
        location = item.get("location") or DEFAULT_LOCATION
        if index > 0:
            time.sleep(REQUEST_DELAY_SEC)
        try:
            data = fetch_serp(api_key, query, location)
        except requests.RequestException as exc:
            results.append({**item, "error": str(exc)})
            print(f"ERROR {query} ({location}): {exc}")
            continue

        organic = summarize_organic(data)
        places = summarize_places(data)
        site_hit = next((row for row in organic if row["is_screenteam"]), None)
        top_competitors = [row for row in organic[:5] if not row["is_screenteam"]]

        payload = {
            **item,
            "intent": infer_intent(query, organic, places),
            "screenteam_rank": site_hit["position"] if site_hit else None,
            "screenteam_result": site_hit,
            "top_competitors": top_competitors,
            "competitor_title_angles": [row.get("title") for row in top_competitors[:3]],
            "competitor_snippet_angles": [row.get("snippet") for row in top_competitors[:3]],
            "map_pack": places[:3],
            "people_also_ask": [
                row.get("question")
                for row in (data.get("peopleAlsoAsk") or [])
                if isinstance(row, dict) and row.get("question")
            ][:3],
            "search_information": data.get("searchInformation"),
        }
        payload["suggested_meta_description"] = suggested_meta_description(payload)
        results.append(payload)
        rank = payload["screenteam_rank"]
        print(f"OK   {query} @ {location} -> rank {rank if rank else 'not in top 10'}")

    SEO.mkdir(parents=True, exist_ok=True)
    report = {
        "generated_at": generated_at,
        "site": f"https://{SITE_DOMAIN}",
        "location_default": DEFAULT_LOCATION,
        "page_count": len(PAGE_QUERY_MAP),
        "query_count": len(results),
        "results": results,
    }
    json_path = SEO / "serp-meta-research.json"
    json_path.write_text(json.dumps(report, indent=2), encoding="utf-8")

    md_lines = [
        "# Screen Team LLC Serper meta research",
        f"Generated: {generated_at}",
        f"Pages mapped: {len(PAGE_QUERY_MAP)}",
        f"Queries run: {len(results)}",
        "",
    ]
    for row in results:
        md_lines.append(f"## {row['query']}")
        md_lines.append(f"- Target page: `{row.get('target_page', '')}`")
        md_lines.append(f"- Location: {row.get('location', DEFAULT_LOCATION)}")
        md_lines.append(f"- GSC impressions: {row.get('gsc_impressions') or 'n/a'}")
        md_lines.append(f"- Intent: {row.get('intent', 'n/a')}")
        md_lines.append(f"- Screen Team rank: {row.get('screenteam_rank') or 'not in top 10'}")
        if row.get("suggested_meta_description"):
            md_lines.append(f"- Suggested meta: {row['suggested_meta_description']}")
        if row.get("top_competitors"):
            md_lines.append("- Top competitors:")
            for comp in row["top_competitors"][:3]:
                md_lines.append(f"  - {comp.get('title')} | {comp.get('domain')}")
        if row.get("people_also_ask"):
            md_lines.append("- People also ask:")
            for question in row["people_also_ask"]:
                md_lines.append(f"  - {question}")
        if row.get("error"):
            md_lines.append(f"- Error: {row['error']}")
        md_lines.append("")

    md_path = SEO / "serp-meta-research.md"
    md_path.write_text("\n".join(md_lines), encoding="utf-8")
    print(f"\nWrote {json_path.relative_to(ROOT)}")
    print(f"Wrote {md_path.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
